import datetime
import os
from google.oauth2 import service_account
from googleapiclient.discovery import build
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from tkinter import Text, filedialog
import tkinter as tk

'''Part 1. Get the date for Monday because that is what the sheet is called'''

def get_previous_monday_date():
    # Get the current date
    today = datetime.datetime.now()

    # Calculate the number of days to subtract to reach the last Monday
    days_to_subtract = today.weekday() % 7

    # Calculate the date of the last Monday
    last_monday = today - datetime.timedelta(days=days_to_subtract)

    # Format the date as a string
    last_monday_str = last_monday.strftime("%Y-%m-%d")

    # print("Date of the last Monday:", last_monday_str)
    return last_monday_str

def find_week_sheet(SERVICE_ACCOUNT_FILE,folder_id,target_file_name):
        
    # Create a service account credentials object
    creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE)

    # Create the Drive API service
    service = build('drive', 'v3', credentials=creds)



    # List all files in the folder
    results = service.files().list(q=f"'{folder_id}' in parents and name = '{target_file_name}'",
                                fields="files(id, name)").execute()
    files = results.get('files', [])

    if files:
        for file in files:
            print(f"Found file: {file['name']} (ID: {file['id']})")
            return file['id']
    else:
        print(f"No file with the name '{target_file_name}' found in the folder.")


'''Part 2. create the excel sheet'''

def create_excel_sheet_skype(CLIENT_CREDS_FILE,FILE_ID,output_file_path):

    # Initialize the Google Sheets API client
    scope = ['https://www.googleapis.com/auth/spreadsheets']
    creds = ServiceAccountCredentials.from_json_keyfile_name(CLIENT_CREDS_FILE, scope)
    client = gspread.authorize(creds)

    # Open the Google Sheet by its file ID
    sheet = client.open_by_key(FILE_ID)

    # Specify the worksheet to work with
    worksheet = sheet.get_worksheet(0)  # Index 0 is the first worksheet

    today= datetime.datetime.now().weekday()
    if today == 4 or today == 5:
        start_row = 117
        end_row = 142
    else:
        start_row = today*28+5
        end_row = start_row + 25


    # Use the Google Sheets API to retrieve cell background colors
    service = build('sheets', 'v4', credentials=creds)
    spreadsheet_id = FILE_ID
    ranges = f'{worksheet.title}!B{start_row}:Z{end_row}'  # Adjust the columns as needed
    request = service.spreadsheets().get(spreadsheetId=spreadsheet_id, ranges=ranges, includeGridData=True)
    response = request.execute()
    cell_data = response['sheets'][0]['data'][0]['rowData']

    # Create a list of lists to store the data and background colors
    data = []
    for i, row in enumerate(cell_data):
        row_data = []
        for j, cell in enumerate(row['values']):
            value = cell.get('formattedValue', '')
            background_color = cell.get('userEnteredFormat', {}).get('backgroundColor', {})
            cell_address = f'{chr(ord("A") + j)}{start_row + i}'
            row_data.append({
                "cell_address": cell_address,
                "value": value,
                "background_color": background_color
            })
        data.append(row_data)

    def rgb_to_hex(rgb):
        return "#{:02x}{:02x}{:02x}".format(int(rgb.get('red', 0) * 255), int(rgb.get('green', 0) * 255), int(rgb.get('blue', 0) * 255))

    # Print the data and background colors as a list of dictionaries
    list_student = []
    for row in data:
        list_time_student = []
        for thing in row:
            if thing['value'] != '':
                value = thing['value']
                color = rgb_to_hex(thing['background_color'])
                list_time_student.append([value, color])
        list_student.append(list_time_student)



    final_list = []
    # print(list_student)

    for itm in list_student:
            # Extract the time value from the first item in the original list
        if itm == []:
            time_value = 'empty'
        else:
            time_value = itm[0][0]

        for name,color in itm[1:]:
            final_list.append([name, color, time_value])





    # Create a DataFrame from the data
    df = pd.DataFrame(final_list, columns=['Student', 'Color', 'Time'])


    # # Define proxy values for colors
    # color_proxy = {
    #     '#00ccff': 'Teacher_1',
    #     '#9fc5e7': 'Teacher_2',
    #     '#d9ead3': 'Teacher_3',
    #     '#fce4cd': 'Teacher_4',
    #     '#ffff00': 'Teacher_5',
    # }

    # # Replace color codes with proxy values
    

        # Get unique colors from the DataFrame
    unique_colors = df['Color'].unique()

    # Create a color proxy dictionary with numeric values
    color_proxy = {color: str(index + 1) for index, color in enumerate(unique_colors)}

    df['Color'] = df['Color'].map(color_proxy)


    # Group by 'Color' and 'Time' and aggregate the students
    grouped = df.groupby(['Color', 'Time'])['Student'].apply(lambda x: ', '.join(x)).reset_index()

    # Create an Excel writer
    with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
        workbook = writer.book

        # Create the 'Output' sheet
        writer.sheets['Output'] = workbook.add_worksheet()

        worksheet = writer.sheets['Output']

        # Add text wrapping for the 'Students' column
        format_wrap = workbook.add_format({'text_wrap': True})
        worksheet.set_column('C:C', None, format_wrap)

        # Write the data to the sheet
        grouped.to_excel(writer, sheet_name='Output', index=False)

    print("Output Excel file has been created.")


'''Part 3. Gets the Appointy info and adds it to the excel sheet'''



def convert(file_path,save_file):
    '''1. convert csv file to excel using pandas'''
    df = pd.read_csv(file_path)
    excel_file = save_file
    df.to_excel(excel_file, index=False)

def delete_bad_columns(file_path):
    ''' Deletes the columns that are uneeded '''

    # Define the column numbers to keep
    columns_to_keep = [2, 5, 18]

    wb = load_workbook(filename=file_path)
    sheet = wb.worksheets[0]
    
    # Calculate the total number of columns in the sheet
    max_column = sheet.max_column

    # Iterate through the columns in reverse order and delete if not in the list of columns to keep
    for col in range(max_column, 0, -1):
        if col not in columns_to_keep:
            sheet.delete_cols(col, 1)

    # Delete the first row (header) if you want to remove it
    sheet.delete_rows(1, 1)

    wb.save(file_path)

def fix_intake(file_path):
    ''' The Students Names are in a weird formatted string it looks like:
    Inake Form:
        Student 1: Name
        Student 2: Name
    This fixes the format and removes everything except for the students names '''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]

    maxr = ws.max_row
    maxc = ws.max_column

    for r in range(1,maxr+1):

        y = ws['C'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace('Student #2 Name (if applicable): ,','')
        y = y.replace('Student #3 Name (if applicable): ','')
        y = y.replace(',\n','')
        y = y.replace('\n','')

        ws['C'+str(r)].value = y

    
    wb.save(file_path)

def no_student_name(file_path):
    '''Some parents do not put in their child's names. In that case this function
    takes in the parents name instead'''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]

    maxr = ws.max_row
    empty = '-'

    for i in range(1,maxr+1):
        student = ws['C' + str(i)].value

        if student == empty:
            parent = ws['B' + str(i)].value
            ws['C' + str(i)].value = parent
    ws.delete_cols(2, 1)
    wb.save(file_path)




def copy_columns(source_file_path, dest_file_path):
    # Load the source workbook
    source_wb = load_workbook(filename=source_file_path)
    
    # Get the first (default) sheet from the source workbook
    source_sheet = source_wb.active

    # Load the destination workbook
    dest_wb = load_workbook(filename=dest_file_path)
    
    # Get the first (default) sheet from the destination workbook
    dest_sheet = dest_wb.active

    # Get the number of rows in the source sheet
    num_rows = source_sheet.max_row
    r = 2

    dest_sheet.cell(row=1, column=5, value='Time From Appointy')
    dest_sheet.cell(row=1, column=6, value='Students From Appointy')

    # Copy the data from the source sheet to the first row of the destination sheet in the 5th and 6th columns
    for row in source_sheet.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=2):
        dest_sheet.cell(row=r, column=5, value=row[0].value)
        dest_sheet.cell(row=r, column=6, value=row[1].value)
        r+=1

    # Save the destination workbook
    dest_wb.save(dest_file_path)

    # Close the workbooks
    source_wb.close()
    dest_wb.close()




def delete_file_excel(file_path):
    # Check if the file exists before attempting to delete it
    if os.path.exists(file_path):
        # Delete the file
        os.remove(file_path)
        print(f"{file_path} has been deleted.")
    else:
        print(f"{file_path} does not exist, so it cannot be deleted.")




def read_config_file(file_path):
    config_dict = {}
    try:
        with open(file_path, 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    key, value = line.split(": ", 1)
                    config_dict[key] = value
        return config_dict
    except FileNotFoundError:
        return "File not found."















# Define a variable to store the selected file path
selected_file_path = ""

# Function to browse for a file
def browse_file():
    global selected_file_path
    selected_file_path = filedialog.askopenfilename()
    log_text.insert(tk.END, f"Selected file: {selected_file_path}\n")

# Define a function to perform the entire process

def process_file():
    

    

    global selected_file_path
    if not selected_file_path:
        log_text.insert(tk.END, "Please select a file first.\n")
        return
    


    log_text.insert(tk.END, "Getting TXT file location\n")

    dictionary = read_config_file('C:/Users/preet/Documents/MATHNASIUM/SCHEDULE_MAKER/functions/read_txt/file_text.txt')
    service_account_key_file = dictionary['service_account_key_file']
    folder_id = dictionary['folder_id']
    output_path= dictionary['output_path']+datetime.date.today().strftime('%Y-%m-%d')+'.xlsx'
    
    log_text.insert(tk.END, "Done!\n")



    
    log_text.insert(tk.END, "Creating file from Google Drive\n")
    FILE_NAME = get_previous_monday_date()
    FILE_ID = find_week_sheet(service_account_key_file,folder_id,FILE_NAME)
    create_excel_sheet_skype(service_account_key_file,FILE_ID,output_path)

    log_text.insert(tk.END, "Done!\n")


    cleaning_file = dictionary['output_path'] + 'output.xlsx'
    
    log_text.insert(tk.END, "Inputing the Students from Appointy\n")
    convert(selected_file_path,cleaning_file)
    delete_bad_columns(cleaning_file)
    fix_intake(cleaning_file)
    no_student_name(cleaning_file)
    copy_columns(cleaning_file, output_path)
    delete_file_excel(cleaning_file)
    log_text.insert(tk.END, "Done!\n")



# Create the main window
root = tk.Tk()
root.title("CREATE LIST OF STUDENTS FOR TODAY")
root.geometry("600x250")

label = tk.Label(root, text="Select a the Appointy file for this day:")
label.pack()
# Create a "Browse" button
browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack()

# Create a Text widget for displaying log messages
log_text = tk.Text(root, height=10, width=60)
log_text.pack()

# Create a button to trigger the entire process
process_button = tk.Button(root, text="Process File", command=process_file)
process_button.pack()

# Start the GUI event loop
root.mainloop()




